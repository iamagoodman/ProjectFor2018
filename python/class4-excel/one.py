# -*- coding: utf-8 -*-
from openpyxl import load_workbook

workbook = load_workbook(filename='test1.xlsx')

print(workbook.sheetnames)

# sheet = workbook['Sheet1']
sheet = workbook.active
print(sheet.dimensions)
cell = sheet['A1']
print(cell.value)

print(sheet.cell(row=2,column=2).value)